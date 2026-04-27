"""
告警 API

处理告警历史查询和导出
"""

import csv
import uuid
from datetime import datetime
from pathlib import Path

from fastapi import APIRouter, Depends, HTTPException, Query
from sqlalchemy.orm import Session

from app.core.database import get_db
from app.core.config import settings
from app.models import AlertConfig
from app.repositories import AlertRepository, VideoSourceRepository, AlertConfigRepository, RegionRepository
from app.schemas.common import ApiResponse
from app.schemas.alert import (
    AlertThresholdGet,
    AlertThresholdUpdate,
    RegionThreshold,
    AlertRecentItem,
    AlertRecentResponse,
    AlertExportResponse,
)

router = APIRouter(prefix="/alerts", tags=["告警管理"])


def _to_threshold_response(config: AlertConfig, regions: list) -> AlertThresholdGet:
    """将数据库配置转换为阈值响应。区域阈值以 Region 字段为准。"""
    region_thresholds = {
        region.region_id: RegionThreshold(
            name=region.name,
            count_warning=region.count_warning,
            count_critical=region.count_critical,
            density_warning=region.density_warning,
            density_critical=region.density_critical,
        )
        for region in regions
    }
    return AlertThresholdGet(
        total_warning_threshold=config.total_warning_threshold,
        total_critical_threshold=config.total_critical_threshold,
        default_region_warning=config.default_region_warning,
        default_region_critical=config.default_region_critical,
        region_thresholds=region_thresholds,
        cooldown_seconds=config.cooldown_seconds,
    )


@router.get("/threshold", response_model=ApiResponse[AlertThresholdGet])
async def get_threshold(
    source_id: str = Query(..., description="数据源 ID"),
    db: Session = Depends(get_db),
):
    """获取预警阈值配置"""
    source_repo = VideoSourceRepository(db)
    if not source_repo.get_by_id(source_id):
        raise HTTPException(status_code=404, detail="数据源不存在")

    config_repo = AlertConfigRepository(db)
    region_repo = RegionRepository(db)
    config = config_repo.get_or_create(source_id=source_id, config_id=str(uuid.uuid4()))
    regions = region_repo.get_by_source_id(source_id)
    return ApiResponse.success(data=_to_threshold_response(config, regions))


@router.post("/threshold", response_model=ApiResponse[AlertThresholdGet])
async def update_threshold(
    request: AlertThresholdUpdate,
    db: Session = Depends(get_db),
):
    """更新预警阈值配置"""
    source_repo = VideoSourceRepository(db)
    if not source_repo.get_by_id(request.source_id):
        raise HTTPException(status_code=404, detail="数据源不存在")

    config_repo = AlertConfigRepository(db)
    region_repo = RegionRepository(db)
    config = config_repo.get_or_create(source_id=request.source_id, config_id=str(uuid.uuid4()))
    regions = {region.region_id: region for region in region_repo.get_by_source_id(request.source_id)}

    updates: dict[str, int | None] = {}
    if request.total_warning_threshold is not None:
        updates["total_warning_threshold"] = request.total_warning_threshold
    if request.total_critical_threshold is not None:
        updates["total_critical_threshold"] = request.total_critical_threshold
    if request.default_region_warning is not None:
        updates["default_region_warning"] = request.default_region_warning
    if request.default_region_critical is not None:
        updates["default_region_critical"] = request.default_region_critical
    if request.cooldown_seconds is not None:
        updates["cooldown_seconds"] = request.cooldown_seconds
    if "region_thresholds" in request.model_fields_set:
        if request.region_thresholds is None:
            for region in regions.values():
                region.count_warning = None
                region.count_critical = None
                region.density_warning = None
                region.density_critical = None
        else:
            for region_id, threshold in request.region_thresholds.items():
                region = regions.get(region_id)
                if region is None:
                    continue
                region.count_warning = threshold.count_warning
                region.count_critical = threshold.count_critical
                region.density_warning = threshold.density_warning
                region.density_critical = threshold.density_critical

    if updates:
        config = config_repo.update(config, **updates)
    if "region_thresholds" in request.model_fields_set:
        db.commit()

    return ApiResponse.success(data=_to_threshold_response(config, list(regions.values())))


@router.get("/recent", response_model=ApiResponse[AlertRecentResponse])
async def get_recent_alerts(
    source_id: str = Query(..., description="数据源 ID"),
    db: Session = Depends(get_db),
):
    """获取最近五次预警"""
    # 验证数据源存在
    source_repo = VideoSourceRepository(db)
    if not source_repo.get_by_id(source_id):
        raise HTTPException(status_code=404, detail="数据源不存在")

    alert_repo = AlertRepository(db)
    alerts = alert_repo.get_by_source_id(source_id, skip=0, limit=5)

    items = [
        AlertRecentItem(
            alert_id=a.alert_id,
            alert_type=a.alert_type,
            level=a.level,
            region_id=a.region_id,
            region_name=a.region_name,
            current_value=int(round(a.current_value)),
            threshold=int(round(a.threshold)),
            timestamp=a.timestamp,
            message=a.message,
        )
        for a in alerts
    ]

    return ApiResponse.success(data=AlertRecentResponse(items=items))


@router.get("/export", response_model=ApiResponse[AlertExportResponse])
async def export_alerts(
    source_id: str = Query(..., description="数据源 ID"),
    from_time: str = Query(..., alias="from", description="开始时间 (ISO 8601)"),
    to_time: str = Query(..., alias="to", description="结束时间 (ISO 8601)"),
    format: str = Query(default="csv", description="导出格式: csv / xlsx"),
    db: Session = Depends(get_db),
):
    """导出告警记录"""
    # 验证数据源存在
    source_repo = VideoSourceRepository(db)
    source = source_repo.get_by_id(source_id)
    if not source:
        raise HTTPException(status_code=404, detail="数据源不存在")

    # 查询告警记录
    alert_repo = AlertRepository(db)
    alerts = alert_repo.get_by_time_range(source_id, from_time, to_time)

    if not alerts:
        raise HTTPException(status_code=404, detail="该时间段内无告警记录")

    # 创建导出目录
    export_dir = Path(settings.BASE_DIR) / "downloads"
    export_dir.mkdir(parents=True, exist_ok=True)

    # 生成文件名
    timestamp = datetime.utcnow().strftime("%Y%m%d_%H%M%S")
    safe_name = "".join(c if c.isalnum() or c in "-_" else "_" for c in source.name)
    filename = f"alerts_{safe_name}_{timestamp}.{format}"
    file_path = export_dir / filename

    # 导出 CSV
    if format == "csv":
        with open(file_path, "w", newline="", encoding="utf-8-sig") as f:
            writer = csv.writer(f)
            writer.writerow(["时间", "类型", "级别", "区域", "当前值", "阈值", "消息"])
            for a in alerts:
                writer.writerow([
                    a.timestamp,
                    a.alert_type,
                    a.level,
                    a.region_name or "",
                    a.current_value,
                    a.threshold,
                    a.message or "",
                ])
    elif format == "xlsx":
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment

            wb = Workbook()
            ws = wb.active
            ws.title = "告警记录"

            headers = ["时间", "类型", "级别", "区域", "当前值", "阈值", "消息"]
            ws.append(headers)

            for a in alerts:
                ws.append([
                    a.timestamp,
                    a.alert_type,
                    a.level,
                    a.region_name or "",
                    a.current_value,
                    a.threshold,
                    a.message or "",
                ])

            # 表头样式
            header_font = Font(bold=True)
            header_fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
            for cell in ws[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center")

            wb.save(file_path)
        except ImportError:
            raise HTTPException(status_code=500, detail="xlsx 导出需要安装 openpyxl")
    else:
        raise HTTPException(status_code=400, detail=f"不支持的格式: {format}")

    logger.info(f"告警记录已导出: {file_path}")

    return ApiResponse.success(data=AlertExportResponse(url=f"/downloads/{filename}"))
