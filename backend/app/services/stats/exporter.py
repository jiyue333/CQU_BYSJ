"""
统计数据导出服务

将统计数据导出为 CSV/Excel 格式
"""

import csv
from datetime import datetime
from io import BytesIO, StringIO
from pathlib import Path
from typing import Optional

from sqlalchemy.orm import Session

from app.core.config import settings
from app.core.logger import logger
from app.repositories.stats_repository import StatsRepository


class StatsExporter:
    """
    统计数据导出服务

    使用方式：
        exporter = StatsExporter(db)
        file_path = exporter.export_csv(source_id, from_time, to_time)
    """

    def __init__(self, db: Session):
        self.db = db
        self.stats_repo = StatsRepository(db)

    def _ensure_export_dir(self) -> Path:
        """确保导出目录存在"""
        export_dir = settings.BASE_DIR / "downloads"
        export_dir.mkdir(parents=True, exist_ok=True)
        return export_dir

    def _generate_filename(self, source_name: str, format: str) -> str:
        """生成导出文件名"""
        timestamp = datetime.utcnow().strftime("%Y%m%d_%H%M%S")
        # 清理文件名中的特殊字符
        safe_name = "".join(c if c.isalnum() or c in "-_" else "_" for c in source_name)
        return f"export_{safe_name}_{timestamp}.{format}"

    def export_csv(
        self,
        source_id: str,
        source_name: str,
        time_from: str,
        time_to: str,
        interval_type: str = "1m",
        region_id: Optional[str] = None,
    ) -> Path:
        """
        导出 CSV 格式

        Args:
            source_id: 数据源 ID
            source_name: 数据源名称（用于文件名）
            time_from: 开始时间
            time_to: 结束时间
            interval_type: 聚合粒度
            region_id: 可选，指定区域

        Returns:
            导出文件路径
        """
        stats = self.stats_repo.get_by_time_range(
            source_id=source_id,
            interval_type=interval_type,
            time_from=time_from,
            time_to=time_to,
            region_id=region_id,
        )

        export_dir = self._ensure_export_dir()
        filename = self._generate_filename(source_name, "csv")
        file_path = export_dir / filename

        with open(file_path, "w", newline="", encoding="utf-8-sig") as f:
            writer = csv.writer(f)

            if region_id:
                # 导出指定区域数据
                writer.writerow(["时间", "区域名称", "平均人数", "最大人数", "最小人数", "拥挤指数"])
                for s in stats:
                    region_data = s.get_region_stats_dict()
                    if region_id in region_data:
                        r = region_data[region_id]
                        writer.writerow([
                            s.time_bucket,
                            r.name,
                            f"{r.avg:.1f}",
                            r.max,
                            r.min,
                            f"{r.crowd_index:.2f}",
                        ])
            else:
                # 导出全局数据
                writer.writerow(["时间", "平均人数", "最大人数", "最小人数", "平均密度", "拥挤指数", "采样数"])
                for s in stats:
                    writer.writerow([
                        s.time_bucket,
                        f"{s.total_count_avg:.1f}",
                        s.total_count_max,
                        s.total_count_min,
                        f"{s.total_density_avg:.4f}",
                        f"{s.crowd_index_avg:.2f}" if s.crowd_index_avg else "",
                        s.sample_count,
                    ])

        logger.info(f"CSV 导出完成: {file_path}")
        return file_path

    def export_xlsx(
        self,
        source_id: str,
        source_name: str,
        time_from: str,
        time_to: str,
        interval_type: str = "1m",
        region_id: Optional[str] = None,
    ) -> Path:
        """
        导出 Excel 格式

        Args:
            source_id: 数据源 ID
            source_name: 数据源名称
            time_from: 开始时间
            time_to: 结束时间
            interval_type: 聚合粒度
            region_id: 可选，指定区域

        Returns:
            导出文件路径

        Raises:
            ImportError: 如果未安装 openpyxl
        """
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment, PatternFill
        except ImportError:
            raise ImportError("导出 Excel 需要安装 openpyxl: pip install openpyxl")

        stats = self.stats_repo.get_by_time_range(
            source_id=source_id,
            interval_type=interval_type,
            time_from=time_from,
            time_to=time_to,
            region_id=region_id,
        )

        export_dir = self._ensure_export_dir()
        filename = self._generate_filename(source_name, "xlsx")
        file_path = export_dir / filename

        wb = Workbook()
        ws = wb.active
        ws.title = "统计数据"

        # 表头样式
        header_font = Font(bold=True)
        header_fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")

        if region_id:
            headers = ["时间", "区域名称", "平均人数", "最大人数", "最小人数", "拥挤指数"]
            ws.append(headers)

            for s in stats:
                region_data = s.get_region_stats_dict()
                if region_id in region_data:
                    r = region_data[region_id]
                    ws.append([
                        s.time_bucket,
                        r.name,
                        round(r.avg, 1),
                        r.max,
                        r.min,
                        round(r.crowd_index, 2),
                    ])
        else:
            headers = ["时间", "平均人数", "最大人数", "最小人数", "平均密度", "拥挤指数", "采样数"]
            ws.append(headers)

            for s in stats:
                ws.append([
                    s.time_bucket,
                    round(s.total_count_avg, 1),
                    s.total_count_max,
                    s.total_count_min,
                    round(s.total_density_avg, 4),
                    round(s.crowd_index_avg, 2) if s.crowd_index_avg else None,
                    s.sample_count,
                ])

        # 应用表头样式
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

        # 调整列宽
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except Exception:
                    pass
            ws.column_dimensions[column_letter].width = max(max_length + 2, 10)

        wb.save(file_path)
        logger.info(f"Excel 导出完成: {file_path}")
        return file_path

    def export_region_stats(
        self,
        source_id: str,
        source_name: str,
        time_from: str,
        time_to: str,
        interval_type: str = "1m",
        format: str = "csv",
    ) -> Path:
        """
        导出所有区域的统计数据（每个区域一个 sheet 或分段）

        Args:
            source_id: 数据源 ID
            source_name: 数据源名称
            time_from: 开始时间
            time_to: 结束时间
            interval_type: 聚合粒度
            format: 导出格式 (csv/xlsx)

        Returns:
            导出文件路径
        """
        stats = self.stats_repo.get_by_time_range(
            source_id=source_id,
            interval_type=interval_type,
            time_from=time_from,
            time_to=time_to,
        )

        if format == "xlsx":
            return self._export_region_stats_xlsx(stats, source_name)
        else:
            return self._export_region_stats_csv(stats, source_name)

    def _export_region_stats_xlsx(self, stats: list, source_name: str) -> Path:
        """导出区域统计到 Excel（每个区域一个 sheet）"""
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment, PatternFill
        except ImportError:
            raise ImportError("导出 Excel 需要安装 openpyxl")

        export_dir = self._ensure_export_dir()
        filename = self._generate_filename(f"{source_name}_regions", "xlsx")
        file_path = export_dir / filename

        wb = Workbook()
        # 删除默认 sheet
        wb.remove(wb.active)

        # 收集所有区域
        all_regions: dict[str, str] = {}  # region_id -> name
        for s in stats:
            for region_id, data in s.get_region_stats_dict().items():
                all_regions[region_id] = data.name

        header_font = Font(bold=True)
        header_fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        headers = ["时间", "平均人数", "最大人数", "最小人数", "拥挤指数"]

        for region_id, region_name in all_regions.items():
            # 创建 sheet（名称最多 31 字符）
            sheet_name = region_name[:31] if region_name else region_id[:31]
            ws = wb.create_sheet(title=sheet_name)
            ws.append(headers)

            for s in stats:
                region_data = s.get_region_stats_dict()
                if region_id in region_data:
                    r = region_data[region_id]
                    ws.append([
                        s.time_bucket,
                        round(r.avg, 1),
                        r.max,
                        r.min,
                        round(r.crowd_index, 2),
                    ])

            # 应用表头样式
            for cell in ws[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center")

        wb.save(file_path)
        logger.info(f"区域统计 Excel 导出完成: {file_path}")
        return file_path

    def _export_region_stats_csv(self, stats: list, source_name: str) -> Path:
        """导出区域统计到 CSV"""
        export_dir = self._ensure_export_dir()
        filename = self._generate_filename(f"{source_name}_regions", "csv")
        file_path = export_dir / filename

        with open(file_path, "w", newline="", encoding="utf-8-sig") as f:
            writer = csv.writer(f)
            writer.writerow(["时间", "区域ID", "区域名称", "平均人数", "最大人数", "最小人数", "拥挤指数"])

            for s in stats:
                for region_id, r in s.get_region_stats_dict().items():
                    writer.writerow([
                        s.time_bucket,
                        region_id,
                        r.name,
                        f"{r.avg:.1f}",
                        r.max,
                        r.min,
                        f"{r.crowd_index:.2f}",
                    ])

        logger.info(f"区域统计 CSV 导出完成: {file_path}")
        return file_path

    def generate_summary_report(
        self,
        source_id: str,
        time_from: str,
        time_to: str,
        interval_type: str = "1h",
    ) -> dict:
        """
        生成统计摘要报告

        Args:
            source_id: 数据源 ID
            time_from: 开始时间
            time_to: 结束时间
            interval_type: 聚合粒度

        Returns:
            统计摘要字典
        """
        stats = self.stats_repo.get_by_time_range(
            source_id=source_id,
            interval_type=interval_type,
            time_from=time_from,
            time_to=time_to,
        )

        if not stats:
            return {
                "period": {"from": time_from, "to": time_to},
                "total_records": 0,
                "summary": None,
            }

        # 全局统计
        all_counts = [s.total_count_avg for s in stats]
        all_densities = [s.total_density_avg for s in stats]
        all_crowd_indices = [s.crowd_index_avg for s in stats if s.crowd_index_avg]

        # 峰值时段
        peak_stat = max(stats, key=lambda s: s.total_count_max)

        # 区域统计汇总
        region_summary: dict[str, dict] = {}
        for s in stats:
            for region_id, r in s.get_region_stats_dict().items():
                if region_id not in region_summary:
                    region_summary[region_id] = {
                        "name": r.name,
                        "counts": [],
                        "crowd_indices": [],
                    }
                region_summary[region_id]["counts"].append(r.avg)
                region_summary[region_id]["crowd_indices"].append(r.crowd_index)

        regions_report = {}
        for region_id, data in region_summary.items():
            counts = data["counts"]
            indices = data["crowd_indices"]
            regions_report[region_id] = {
                "name": data["name"],
                "avg_count": sum(counts) / len(counts) if counts else 0,
                "max_count": max(counts) if counts else 0,
                "avg_crowd_index": sum(indices) / len(indices) if indices else 0,
            }

        return {
            "period": {"from": time_from, "to": time_to},
            "total_records": len(stats),
            "summary": {
                "avg_count": sum(all_counts) / len(all_counts),
                "max_count": max(s.total_count_max for s in stats),
                "min_count": min(s.total_count_min for s in stats),
                "avg_density": sum(all_densities) / len(all_densities),
                "avg_crowd_index": sum(all_crowd_indices) / len(all_crowd_indices) if all_crowd_indices else None,
                "peak_time": peak_stat.time_bucket,
                "peak_count": peak_stat.total_count_max,
            },
            "regions": regions_report,
        }
